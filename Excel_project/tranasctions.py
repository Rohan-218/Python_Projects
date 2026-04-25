import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['a1']
    # cell=sheet.cell(1,1)
    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        correct_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = correct_price

    values = Reference(
            sheet ,
            min_row=2 ,
            min_col =4,
            max_row = sheet.max_row,
            max_col=4
            )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e9')
    wb.save(filename)
    print("\nExecuted without Errors.......")


filename = input("Enter the name of the excel file : ")
process_workbook(filename)